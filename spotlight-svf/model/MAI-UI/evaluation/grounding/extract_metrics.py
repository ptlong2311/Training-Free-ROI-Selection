#!/usr/bin/env python3

import json
import os
import csv
import argparse
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd

def extract_action_acc_from_json(json_file_path: str) -> Optional[Dict]:
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if 'metrics' in data and 'overall' in data['metrics'] and 'action_acc' in data['metrics']['overall']:
            action_acc = data['metrics']['overall']['action_acc']
            num_correct = data['metrics']['overall'].get('num_correct_action', 0)
            num_total = data['metrics']['overall'].get('num_total', 0)
            wrong_format_num = data['metrics']['overall'].get('wrong_format_num', 0)
            
            return {
                'filename': os.path.basename(json_file_path).replace('.json', ''),
                'action_acc': action_acc,
                'num_correct_action': num_correct,
                'num_total': num_total,
                'wrong_format_num': wrong_format_num,
                'file_path': json_file_path
            }
        else:
            print(f"Warning: metrics/overall/action_acc path not found in {json_file_path}")
            return None
            
    except json.JSONDecodeError as e:
        print(f"Error: Cannot parse JSON file {json_file_path}: {e}")
        return None
    except FileNotFoundError:
        print(f"Error: File not found {json_file_path}")
        return None
    except Exception as e:
        print(f"Error: Unknown error processing file {json_file_path}: {e}")
        return None

def find_checkpoint_folders(directory: str) -> List[str]:
    checkpoint_folders = []
    directory_path = Path(directory)
    
    if not directory_path.exists():
        print(f"Error: Directory not found {directory}")
        return []
    
    if not directory_path.is_dir():
        print(f"Error: {directory} is not a directory")
        return []
    
    for folder in directory_path.iterdir():
        if folder.is_dir() and folder.name.startswith("checkpoint"):
            checkpoint_folders.append(str(folder))
    
    return sorted(checkpoint_folders)

def find_json_files(directory: str) -> List[str]:
    json_files = []
    directory_path = Path(directory)
    
    if not directory_path.exists():
        print(f"Error: Directory not found {directory}")
        return []
    
    if not directory_path.is_dir():
        print(f"Error: {directory} is not a directory")
        return []
    
    for json_file in directory_path.glob("*.json"):
        json_files.append(str(json_file))
    
    return sorted(json_files)

def extract_metrics_from_checkpoints(input_directory: str, output_file: str = None) -> None:
    checkpoint_folders = find_checkpoint_folders(input_directory)
    
    if not checkpoint_folders:
        print(f"No checkpoint folders found in directory {input_directory}")
        return
    
    print(f"Found {len(checkpoint_folders)} checkpoint folders")
    
    all_metrics_data = {}
    
    for checkpoint_folder in checkpoint_folders:
        folder_name = os.path.basename(checkpoint_folder)
        print(f"\nProcessing folder: {folder_name}")
        
        json_files = find_json_files(checkpoint_folder)
        if not json_files:
            print(f"  Warning: No JSON files found in {folder_name}")
            continue
        
        print(f"  Found {len(json_files)} JSON files")
        
        folder_metrics = {}
        for json_file in json_files:
            filename = os.path.basename(json_file).replace('.json', '')
            print(f"  Processing file: {filename}")
            metric = extract_action_acc_from_json(json_file)
            if metric:
                folder_metrics[filename] = round(metric['action_acc'], 4)
        
        if folder_metrics:
            all_metrics_data[folder_name] = folder_metrics
    
    if not all_metrics_data:
        print("No valid metrics data extracted")
        return
    
    df = pd.DataFrame(all_metrics_data)
    df = df.fillna(0)
    df = df.round(4)
    
    if output_file is None:
        output_file = os.path.join(input_directory, "metrics_comparison.xlsx")
    elif not output_file.endswith(('.xlsx', '.csv')):
        output_file += '.xlsx'
    
    try:
        if output_file.endswith('.xlsx'):
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Metrics Comparison', index_label='Dataset')
                
                worksheet = writer.sheets['Metrics Comparison']
                
                from openpyxl.styles import PatternFill
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                
                for row_idx in range(2, len(df) + 2):
                    row_values = []
                    for col_idx in range(2, len(df.columns) + 2):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None:
                            row_values.append((col_idx, float(cell_value)))
                    
                    if row_values:
                        max_col = max(row_values, key=lambda x: x[1])[0]
                        worksheet.cell(row=row_idx, column=max_col).fill = red_fill
        else:
            df.to_csv(output_file, encoding='utf-8')
        
        print(f"\nSuccessfully extracted metrics and saved to: {output_file}")
        
        print("\n=== Metrics Comparison Summary ===")
        print(f"{'Dataset':<20}", end="")
        for checkpoint in df.columns:
            print(f"{checkpoint:<15}", end="")
        print()
        print("-" * (20 + 15 * len(df.columns)))
        
        for dataset in df.index:
            print(f"{dataset:<20}", end="")
            row_values = df.loc[dataset]
            max_value = row_values.max()
            for checkpoint in df.columns:
                value = row_values[checkpoint]
                if value == max_value and value > 0:
                    print(f"*{value:.4f}*{'':<8}", end="")
                else:
                    print(f"{value:.4f}{'':<10}", end="")
            print()
        
        print("-" * (20 + 15 * len(df.columns)))
        print("Note: *Marks* indicate the highest score for that dataset")
        
    except Exception as e:
        print(f"Error: Cannot write to file {output_file}: {e}")
        import traceback
        traceback.print_exc()

def extract_metrics_to_csv(input_directory: str, output_csv: str = None) -> None:
    checkpoint_folders = find_checkpoint_folders(input_directory)
    
    if checkpoint_folders:
        print("Checkpoint folders detected, using new processing method...")
        extract_metrics_from_checkpoints(input_directory, output_csv)
        return
    
    json_files = find_json_files(input_directory)
    
    if not json_files:
        print(f"No JSON files found in directory {input_directory}")
        return
    
    print(f"Found {len(json_files)} JSON files")
    
    metrics_data = []
    for json_file in json_files:
        print(f"Processing file: {os.path.basename(json_file)}")
        metric = extract_action_acc_from_json(json_file)
        if metric:
            metric['action_acc'] = round(metric['action_acc'], 4)
            metrics_data.append(metric)
    
    if not metrics_data:
        print("No valid metrics data extracted")
        return
    
    if output_csv is None:
        output_csv = os.path.join(input_directory, "metrics_summary.csv")
    
    try:
        with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['filename', 'action_acc', 'num_correct_action', 'num_total', 'wrong_format_num', 'file_path']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for metric in metrics_data:
                writer.writerow(metric)
        
        print(f"\nSuccessfully extracted {len(metrics_data)} metrics and saved to: {output_csv}")
        
        print("\n=== Metrics Summary ===")
        print(f"{'Filename':<20} {'Action Acc':<15} {'Correct/Total':<15}")
        print("-" * 50)
        for metric in sorted(metrics_data, key=lambda x: x['action_acc'], reverse=True):
            accuracy_str = f"{metric['action_acc']:.4f}"
            count_str = f"{metric['num_correct_action']}/{metric['num_total']}"
            print(f"{metric['filename']:<20} {accuracy_str:<15} {count_str:<15}")
        
        total_correct = sum(m['num_correct_action'] for m in metrics_data)
        total_samples = sum(m['num_total'] for m in metrics_data)
        avg_accuracy = round(total_correct / total_samples, 4) if total_samples > 0 else 0
        
        print("-" * 50)
        avg_accuracy_str = f"{avg_accuracy:.4f}"
        print(f"{'Overall Average':<20} {avg_accuracy_str:<15} {total_correct}/{total_samples}")
        
    except Exception as e:
        print(f"Error: Cannot write CSV file {output_csv}: {e}")

def main():
    parser = argparse.ArgumentParser(
        description="Extract key metrics from JSON evaluation files and organize into Excel/CSV, supporting multi-checkpoint comparison.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Usage Examples:
  # Basic usage: Process a directory containing checkpoint folders
  python extract_metrics.py --input /path/to/metrics_output
  python extract_metrics.py -i /path/to/metrics_output -o comparison.xlsx
  
  # Using positional arguments (simplified)
  python extract_metrics.py /path/to/metrics_output
  python extract_metrics.py /path/to/metrics_output --output results.xlsx
  
  # Process a single folder (backward compatible)
  python extract_metrics.py --input /path/to/single_folder --output summary.csv
  
  # Interactive mode (run without arguments)
  python extract_metrics.py
        """
    )
    
    parser.add_argument(
        'input_directory',
        nargs='?',
        help='Input directory path (containing checkpoint folders or JSON evaluation files)'
    )
    
    parser.add_argument(
        '--input', '-i',
        dest='input_path',
        metavar='DIR',
        help='Input directory path (containing checkpoint folders or JSON evaluation files)'
    )
    
    parser.add_argument(
        '--output', '-o',
        dest='output_path',
        metavar='FILE',
        help='Output file path (.xlsx or .csv). Default: metrics_comparison.xlsx for multi-checkpoint, metrics_summary.csv for single folder'
    )
    
    parser.add_argument(
        '--format', '-f',
        dest='output_format',
        choices=['xlsx', 'csv', 'auto'],
        default='auto',
        help='Output format: xlsx (Excel), csv, or auto (selects based on input). Default: auto'
    )

    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Show verbose processing information'
    )
    
    args = parser.parse_args()
    
    input_dir = args.input_directory or args.input_path
    
    if not input_dir:
        print("Metrics Extraction Tool")
        print("=" * 40)
        print("Supports multi-checkpoint folder comparison analysis")
        print("=" * 40)
        
        input_dir = input("Please enter directory path: ").strip()
        if not input_dir:
            print("Error: Input directory path is required")
            return

        output_path = input("Please enter output file path (enter for default): ").strip()
        output_file = output_path if output_path else None
        format_input = input("Please select output format [xlsx/csv/auto] (default auto): ").strip().lower()
        output_format = format_input if format_input in ['xlsx', 'csv', 'auto'] else 'auto'
    else:
        output_file = args.output_path
        output_format = args.output_format
    
    if not os.path.exists(input_dir):
        print(f"Error: Input directory does not exist: {input_dir}")
        return
    
    if not os.path.isdir(input_dir):
        print(f"Error: Input path is not a directory: {input_dir}")
        return
    
    if output_file and output_format != 'auto':
        if output_format == 'xlsx' and not output_file.endswith('.xlsx'):
            if output_file.endswith('.csv'):
                output_file = output_file[:-4] + '.xlsx'
            else:
                output_file += '.xlsx'
        elif output_format == 'csv' and not output_file.endswith('.csv'):
            if output_file.endswith('.xlsx'):
                output_file = output_file[:-5] + '.csv'
            else:
                output_file += '.csv'
    if args.verbose:
        print(f"Input directory: {input_dir}")
        print(f"Output file: {output_file or 'Default'}")
        print(f"Output format: {output_format}")
        print("-" * 40)
    try:
        extract_metrics_to_csv(input_dir, output_file)
    except Exception as e:
        print(f"Error: Exception occurred during processing: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    main()